import openpyxl as op

subcategories_dict = {}

filename = 'Бланк заказа.xlsx'

wb = op.load_workbook(filename, data_only=True)
sheet = wb.active

max_rows = sheet.max_row

for i in range(7, max_rows + 1):
    article = sheet.cell(row=i,column=2).value
    subcategory = sheet.cell(row=i,column=12).value
    if not article:
        continue
    if subcategory not in subcategories_dict:
        subcategories_dict[subcategory] = [article]
    else:
        subcategories_dict[subcategory].append(article)

sorteddict = dict(sorted(subcategories_dict.items()))
    
with open('subcategories.txt', 'w') as myfile:
    
    for key, value in sorteddict.items():
        string_values = ', '.join(value)
        string_to_write = key + ' = ' + string_values + '\n'
        myfile.write(string_to_write)
